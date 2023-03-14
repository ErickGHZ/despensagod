import openpyxl  #  Importar la librería openpyxl 
from tabulate import tabulate  #  Importar la función tabulate de la librería tabulate
libro = openpyxl.load_workbook('ventas.xlsx')  #  Cargar el archivo Excel 'ventas.xlsx' y asignar el libro a la variable libro
hoja = libro['Productos']  # Seleccionar la hoja 'Productos' del libro y asignarla a la variable hoja


class Productos:  # Clase Productos
    def __init__(self): # Constructor de la clase Productos
        pass # Inicializa el objeto de la clase Productos

    def listarProductos(self) -> bool: # Metodo para listar los productos registrados
        datos = []  #  Crear una lista vacía para almacenar los datos de la hoja
        for fila in hoja.iter_rows():  #  Recorrer todas las filas de la hoja
            fila_datos = []  #  Recorrer todas las filas de la hoja
            for celda in fila:   #  Recorrer todas las celdas de la fila actual
                fila_datos.append(celda.value)  #  Agregar su valor a la lista fila_datos
            datos.append(fila_datos)  #  Agregar la lista fila_datos a la lista datos
        print(tabulate(datos, headers='firstrow', tablefmt='fancy_grid')) # Imprimir los datos en forma de tabla utilizando la librería tabulate, el argumento "headers" especifica que la primera fila debe ser utilizada como encabezados, el argumento "tablefmt" especifica el formato de la tabla 
        return False # Regresa False si ocurrio un error en el metodo

    def insertarProducto(self, sku: str, nombre: str, unidad: str) -> bool: # Metodo para insertar un producto
        # TODO programar el método insertarProducto()
        return False # Regresa False si ocurrio un error en el metodo

    def buscarProducto(self) -> bool: # Metodo para buscar un producto
        SKU = input("Ingresa el codigo SKU: ") # Solicita al usuario que ingrese el código SKU del producto a buscar y lo almacena en la variable "SKU"
        while SKU == "": # Ciclo "while" que se ejecuta mientras "SKU" esté vacía (es decir, mientras el usuario no haya ingresado ningún valor)
            SKU = input("Ingresa el codigo SKU: ") # Solicita al usuario que ingrese el código SKU del producto a buscar
        SKU_1 = int(SKU) # Convierte la cadena "SKU" en un número entero y lo almacena en la variable "SKU_1"
        for row in hoja.iter_rows(min_row=2, values_only=True): # Ciclo "for" que itera sobre todas las filas de una hoja de cálculo (representada por el objeto "hoja") a partir de la fila 2. El parámetro "values_only=True" indica que solo se deben tomar en cuenta los valores de las celdas (y no las celdas en sí).
            if row[0] == SKU_1: # Si el primer valor de la fila coincide con el código SKU buscado ("SKU_1")
                i = 0
                while i <= 2: # Ciclo "while" que itera sobre los primeros tres valores de la fila (usando la variable "i")
                    print(row[i]) # Imprime cada valor de la fila
                    i += 1 # Incrementa la variable "i" en 1
                break # Detiene el ciclo "for" mediante la instrucción "break"
        else: # Si el ciclo "for" se ejecuta completamente sin encontrar una coincidencia
            print("Producto no registrado") # Imprime un mensaje indicando que el producto no está registrado
        return False # Retorna "False" si ocurrió un error en el método.

    def borrarProducto(self, sku: str) -> bool: # Metodo para borrar un producto
        # TODO programar el método borrarProducto()
        return False # Regresa False si ocurrio un error en el metodo

    def actualizarProducto(self, sku: str, nombre: str, unidad: str) -> bool: # Metodo para actualizar un producto
        # TODO programar el método actualizarProducto()
        return False # Regresa False si ocurrio un error en el metodo

productos = Productos()
productos.listarProductos()
